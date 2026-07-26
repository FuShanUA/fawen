"""Report generation: conclusion extraction + Excel summary.

extract_conclusion parses audit report text to derive structured
conclusion fields. generate_excel builds a styled multi-sheet workbook
with hyperlinks to audit reports and original PDFs.
"""

import os
import re

from .enterprise import (
    get_enterprise_note,
    get_enterprise_inst,
    find_source_dir,
    derive_run_paths,
)


def extract_conclusion(report_text: str):
    """Parse audit report text → (audit_conclusion, final_conclusion, detail).

    Tries explicit **审计结论** / **最终结论** markers first,
    then falls back to keyword scanning of the conclusion section.
    """
    if not report_text:
        return '一般性问题', '处理失败', '处理失败，请检查日志'

    audit_conc = None
    final_conc = None

    # Pattern 1: explicit two-level conclusion
    m_audit = re.search(
        r'\*\*审计结论\*\*[：:]\s*\[?(踩红线（一票否决）|踩红线|红线质疑|一般性问题|无)[\]】]?',
        report_text,
    )
    m_final = re.search(
        r'\*\*最终结论\*\*[：:]\s*\[?(不通过|复审|修改|通过)[\]】]?',
        report_text,
    )
    if m_audit:
        audit_conc = m_audit.group(1)
        if audit_conc == '踩红线':
            audit_conc = '踩红线（一票否决）'
    if m_final:
        final_conc = m_final.group(1)

    # Fallback: scan conclusion section
    if not audit_conc:
        conclusion_section = report_text
        found_marker = False
        for marker in ['最终审计判定', '审计判定结论', '判定结论', '审计结论',
                       '总体判定结论', '最终判定', '总体结论', '综合判定结论',
                       '综合判定', '最终审计结论', '最终结论']:
            idx = report_text.rfind(marker)
            if idx >= 0:
                conclusion_section = report_text[idx:]
                found_marker = True
                break
        if not found_marker:
            conclusion_section = report_text[-600:]

        absolute_patterns = r'张冠李戴|模板.*?套用|其他企业.*?名称|其他行业|纯模板'
        has_absolute = False
        for m in re.finditer(absolute_patterns, conclusion_section):
            neg_prefix = conclusion_section[max(0, m.start() - 3):m.start()]
            if any(neg in neg_prefix for neg in ['未', '无', '不', '没有']):
                continue
            has_absolute = True
            break

        if re.search(r'demo|占位符|虚构.*?系统|截图.*?造假', conclusion_section, re.IGNORECASE):
            for m in re.finditer(r'demo|占位符|虚构.*?系统|截图.*?造假', conclusion_section, re.IGNORECASE):
                neg_prefix = conclusion_section[max(0, m.start() - 3):m.start()]
                if not any(neg in neg_prefix for neg in ['未', '无', '不', '没有', '非']):
                    has_absolute = True
                    break

        if re.search(r'文控.*?冒充|数据管理文件.*?充当.*?数据管理办法|概念.*?混淆', conclusion_section):
            has_absolute = True
        if re.search(r'合同.*?纯度.*?不足|无.*?数据管理.*?合同|合同.*?(单一|集中|系统集成)', conclusion_section):
            has_absolute = True
        if re.search(r'数据质量.*?(完全|全部|所有).*?(无|没有|缺失).*?截图', conclusion_section):
            has_absolute = True

        if has_absolute:
            audit_conc = '踩红线（一票否决）'
        elif re.search(r'红线质疑|重大质疑|重大整改', conclusion_section):
            audit_conc = '红线质疑'
        elif re.search(r'一般性|一般.*?问题|一般.*?缺陷', conclusion_section):
            audit_conc = '一般性问题'
        elif re.search(r'通过', conclusion_section[-200:]):
            audit_conc = '无'
        else:
            audit_conc = '红线质疑'

    # Derive final conclusion if not explicit
    if not final_conc:
        if audit_conc == '踩红线（一票否决）':
            final_conc = '不通过'
        elif audit_conc == '红线质疑':
            suspicion_count = len(re.findall(
                r'红线质疑|重大质疑',
                conclusion_section if 'conclusion_section' in dir() else report_text[-1000:],
            ))
            final_conc = '不通过' if suspicion_count >= 3 else ('复审' if suspicion_count >= 1 else '修改')
        elif audit_conc == '一般性问题':
            final_conc = '修改'
        else:
            final_conc = '通过'

    # Extract detail (first substantive line)
    detail = ''
    skip_kw = ['总耗时', '阶段:', 'tokens', 'token', 'phase', 'extract',
               'consistency', 'glm', 'qwen', 'gcs', 'prompt', 'completion',
               'reasoning', 'dcmm 审计报告', '---', '基于', '根据', '依据',
               '以下', '审计准则', '阶段一', '阶段二', '专家级', '质量标准',
               '专家沉淀', '视觉分析结果', '对 <', '对企业', '的最终',
               '审计判定如下', '判定如下', '系统截图', '文本审计结论',
               '请直接输出', '审计结论', '最终结论']
    red_kw = ['踩红线', '红线', '一票否决', '否决', '造假', 'demo', '冒充',
              '张冠李戴', '模板', '违规', '缺失', '不予通过', '不通过',
              '严重', '质疑']
    report_body = report_text.split('---', 1)[1] if '---' in report_text else report_text
    if audit_conc == '踩红线（一票否决）':
        for line in report_body.split('\n'):
            s = line.strip()
            if len(s) < 30 or s.startswith('#') or s.startswith('**'):
                continue
            if any(kw in s.lower() for kw in skip_kw):
                continue
            if any(kw in s for kw in red_kw):
                detail = s[:200]
                break
    if not detail:
        for line in report_body.split('\n'):
            s = line.strip()
            if len(s) < 30:
                continue
            if s.startswith('#') or (s.startswith('**') and len(s) < 50):
                continue
            if any(kw in s.lower() for kw in skip_kw):
                continue
            if s.startswith('###') or s.startswith('####'):
                continue
            detail = s[:200]
            break

    return audit_conc, final_conc, detail


def generate_excel(results: list, total_time: float, batch_dirs=None,
                   out_dir: str = None, excel_name: str = None,
                   base_dir: str = None):
    """Build a styled multi-sheet Excel summary.

    Parameters:
        results:    list of result dicts from the engine
        total_time: wall-clock seconds for the whole batch
        batch_dirs: list of batch directory names (for path derivation)
        out_dir:    output directory for .md reports
        excel_name: filename for the Excel
        base_dir:   base audit data directory
    """
    from openpyxl import Workbook
    from openpyxl import load_workbook as _load_wb
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.comments import Comment
    from openpyxl.worksheet.datavalidation import DataValidation

    if batch_dirs is None:
        batch_dirs = ['2、三级（第一天）']
    if out_dir is None:
        out_dir = '.'
    if base_dir is None:
        base_dir = os.path.dirname(out_dir)

    _default_out, _default_excel = derive_run_paths(batch_dirs, out_dir)
    if excel_name is None:
        excel_name = _default_excel

    headers = ["#", "评估机构", "企业名称", "目标级别", "企业类型",
               "审计结论", "最终结论", "结论详情", "底稿链接", "报告链接", "备注"]
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    def style_header(ws):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
        widths = [6, 20, 30, 8, 8, 16, 10, 40, 10, 10, 12]
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = w
        ws.row_dimensions[1].height = 30
        ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}1'
        ws.freeze_panes = 'A2'
        dv_audit = DataValidation(
            type='list',
            formula1='"踩红线（一票否决）,红线质疑,一般性问题,无"',
            allow_blank=True,
        )
        dv_audit.add('F2:F10000')
        ws.add_data_validation(dv_audit)
        dv_final = DataValidation(
            type='list', formula1='"不通过,复审,修改,通过"', allow_blank=True,
        )
        dv_final.add('G2:G10000')
        ws.add_data_validation(dv_final)

    # Build institution mapping from folder structure
    _num_to_inst = {}
    for batch_dir_name in batch_dirs:
        bdir = (batch_dir_name if os.path.isabs(batch_dir_name)
                else os.path.join(base_dir, batch_dir_name))
        if not os.path.exists(bdir):
            continue
        for inst_dir_name in os.listdir(bdir):
            inst_dir_path = os.path.join(bdir, inst_dir_name)
            if not os.path.isdir(inst_dir_path) or inst_dir_name.startswith('.'):
                continue
            inst_clean = re.sub(r'^\d+、', '', inst_dir_name)
            for sub_name in os.listdir(inst_dir_path):
                if not os.path.isdir(os.path.join(inst_dir_path, sub_name)):
                    continue
                m = re.match(r'(\d+)\s*[、,，]', sub_name)
                if m:
                    _num_to_inst[m.group(1)] = inst_clean

    # Load existing Excel for incremental merge
    excel_path = os.path.join(os.path.dirname(out_dir), excel_name)
    existing_data = {}
    if os.path.exists(excel_path):
        try:
            old_wb = _load_wb(excel_path)
            if '审计汇总' in old_wb.sheetnames:
                old_ws = old_wb['审计汇总']
                for row in old_ws.iter_rows(min_row=2, max_row=old_ws.max_row, values_only=True):
                    if row[2]:
                        existing_data[str(row[2]).strip()] = {
                            'num': row[0], 'inst': row[1], 'name': str(row[2]).strip(),
                            'level': row[3], 'type': row[4], 'conclusion': row[5],
                            'detail_conc': row[6], 'detail': row[7],
                            'note': row[10] if len(row) > 10 else '',
                        }
            old_wb.close()
            print(f'Loaded {len(existing_data)} existing enterprises from Excel')
        except Exception as e:
            print(f'Warning: could not load existing Excel: {e}')

    # Build data rows
    rows = []
    for r in results:
        report = ''
        if r.get('report_path') and os.path.exists(r['report_path']):
            with open(r['report_path'], 'r') as f:
                report = f.read()
        num = re.match(r'(\d+)', r['name'])
        num = num.group(1) if num else ''
        level = '3级' if '3级' in r['name'] or '三级' in r['name'] else ('稳健级' if '稳健级' in r['name'] else '')
        ent_type = '乙方' if '乙方' in r['name'] else '甲方'
        clean_name = re.sub(r'^\d+\s*[、,，_]\s*', '', r['name'])
        for marker in ['3级甲方', '3级乙方', '三级甲方', '三级乙方', '稳健级甲方',
                       '稳健级乙方', '3级', '三级', '稳健级', '甲方', '乙方',
                       '工联数据', 'PDF', 'pdf']:
            clean_name = clean_name.replace(marker, '')
        for sep in ['+', '-', '—', '－', '--', '  ']:
            clean_name = clean_name.replace(sep, ' ')
        clean_name = re.sub(r'^[、,，_\s]+', '', clean_name)
        clean_name = re.sub(r'\s+[甲乙]$', '', clean_name)
        clean_name = re.sub(r'\s+三$', '', clean_name)
        clean_name = re.sub(r'_+$', '', clean_name)
        clean_name = clean_name.strip()
        inst = _num_to_inst.get(num, '') or get_enterprise_inst(r['name'])
        note = get_enterprise_note(r['name'])
        audit_conc, final_conc, detail = extract_conclusion(report)
        if r.get('error'):
            audit_conc = '一般性问题'
            final_conc = '处理失败'
            detail = r['error'][:200]
        rows.append({
            'num': num, 'inst': inst, 'name': clean_name, 'level': level,
            'type': ent_type, 'conclusion': audit_conc, 'detail_conc': final_conc,
            'detail': detail, 'note': note, 'error': r.get('error'),
            'report_path': r.get('report_path', ''), 'time': r.get('timing', {}).get('total', 0),
        })

    # Merge existing entries not in new results
    new_nums = set(r.get('num', '') for r in rows if r.get('num'))
    new_name_keys = set(re.sub(r'[^\u4e00-\u9fff]', '', r['name'])[:10] for r in rows if r.get('name'))
    for ent_name, ent_data in existing_data.items():
        ent_num = str(ent_data.get('num', ''))
        ent_key = re.sub(r'[^\u4e00-\u9fff]', '', ent_name)[:10]
        if ent_num in new_nums or ent_key in new_name_keys:
            continue
        rows.append({
            'num': ent_data.get('num', ''), 'inst': ent_data.get('inst', ''),
            'name': ent_data['name'], 'level': ent_data.get('level', ''),
            'type': ent_data.get('type', ''), 'conclusion': ent_data.get('conclusion', ''),
            'detail_conc': ent_data.get('detail_conc', ''), 'detail': ent_data.get('detail', ''),
            'note': ent_data.get('note', ''), 'error': None,
            'report_path': '', 'time': 0,
        })

    # Reconstruct report_path and source_link
    for r in rows:
        if not r.get('report_path'):
            safe_name = r['name'][:20].replace('/', '_').replace(' ', '_')
            candidate = os.path.join(out_dir, f'{safe_name}.md')
            if os.path.exists(candidate):
                r['report_path'] = candidate
        _, source_link = find_source_dir(
            r.get('name', ''), r.get('num', ''),
            batch_dirs=batch_dirs, base_dir=base_dir,
        )
        r['source_link'] = source_link or ''

    rows.sort(key=lambda r: int(r['num']) if r['num'].isdigit() else 9999)

    wb = Workbook()
    ws_master = wb.active
    ws_master.title = "审计汇总"
    style_header(ws_master)

    def fill_row(ws, row_idx, r):
        data = [r['num'], r['inst'], r['name'], r['level'], r['type'],
                r['conclusion'], r['detail_conc'], r['detail'],
                r['report_path'] if not r.get('error') and r['report_path'] else '',
                '' if r.get('error') else f'file://{out_dir}',
                r['note']]
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            if col == 6:
                if '踩红线' in str(val): cell.fill = red_fill
                elif '质疑' in str(val): cell.fill = yellow_fill
            if col == 7:
                if str(val) == '不通过': cell.fill = red_fill
                elif str(val) == '复审': cell.fill = yellow_fill
            if col == 9 and r['report_path'] and not r.get('error'):
                cell.value = '🔗 查看底稿'
                cell.hyperlink = r['report_path']
                cell.comment = Comment(r['report_path'], 'audit')
                cell.font = Font(color='0563C1', underline='single')
            if col == 10 and not r.get('error'):
                cell.value = '📄 完整报告'
                _src = r.get('source_link', '') or r['report_path']
                cell.hyperlink = _src
                cell.comment = Comment(_src, 'audit')
                cell.font = Font(color='0563C1', underline='single')

    for row_idx, r in enumerate(rows, 2):
        fill_row(ws_master, row_idx, r)

    # Per-institution tabs
    inst_groups = {}
    for r in rows:
        inst = r['inst'] or '未知机构'
        if inst not in inst_groups:
            inst_groups[inst] = []
        inst_groups[inst].append(r)
    for inst, inst_rows in inst_groups.items():
        safe = inst[:20].replace('/', '_')
        ws = wb.create_sheet(title=safe)
        style_header(ws)
        for row_idx, r in enumerate(inst_rows, 2):
            fill_row(ws, row_idx, r)

    out_path = os.path.join(os.path.dirname(out_dir), excel_name)
    wb.save(out_path)
    print(f'Excel saved: {out_path} ({len(rows)} rows [{len(results)} new + {len(rows) - len(results)} existing], {len(inst_groups) + 1} sheets)')
    return out_path
