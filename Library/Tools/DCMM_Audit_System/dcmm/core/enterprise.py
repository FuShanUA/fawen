"""Enterprise list loading, name resolution, and path derivation.

Loads the enterprise list Excel to resolve notes (备注) and assessment
institution (评估机构) for each enterprise. Also derives output paths
based on batch directory names.
"""

import os
import re

_enterprise_notes = {}
_enterprise_inst = {}
_loaded = False


def load_enterprise_list(path: str):
    """Load the enterprise list Excel into module-level caches."""
    global _loaded, _enterprise_notes, _enterprise_inst
    if _loaded:
        return
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True)
        for sheet_name in ['第一天三级', '第二天三级']:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            headers = [cell.value for cell in ws[1]]
            ent_col = headers.index('企业名称') + 1
            note_col = headers.index('备注') + 1
            inst_col = headers.index('评估机构') + 1
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
                ent_name = row[ent_col - 1]
                note = row[note_col - 1] if row[note_col - 1] else ''
                inst = row[inst_col - 1] if row[inst_col - 1] else ''
                if ent_name:
                    _enterprise_notes[str(ent_name).strip()] = str(note).strip()
                    _enterprise_inst[str(ent_name).strip()] = str(inst).strip()
    except Exception as e:
        print(f'Warning: could not load enterprise list: {e}')
    _loaded = True


def get_enterprise_note(ent_name: str) -> str:
    clean = re.sub(r'\d+级|甲方|乙方|[+－\-]', '', ent_name).strip()
    for key, note in _enterprise_notes.items():
        if clean in key or key in clean:
            return note
    return ''


def get_enterprise_inst(ent_name: str) -> str:
    clean = re.sub(r'\d+级|甲方|乙方|[+－]', '', ent_name).strip()
    for key, inst in _enterprise_inst.items():
        if clean in key or key in clean:
            return inst
    return ''


def derive_run_paths(batch_dirs: list, default_out_dir: str):
    """Derive output dir and Excel filename suffix from batch_dirs.

    第一天 → no suffix (审计结果/DCMM审计汇总.xlsx)
    第二天 → _第二天 suffix
    Returns (out_dir, excel_name).
    """
    suffix = ''
    if batch_dirs:
        for bd in batch_dirs:
            base = os.path.basename(bd.rstrip('/'))
            m = re.search(r'第[一二三四五六七八九十]+天', base)
            if m and m.group() != '第一天':
                suffix = '_' + m.group()
                break
    if suffix:
        parent = os.path.dirname(default_out_dir)
        out_dir = os.path.join(parent, '审计结果' + suffix)
        excel_name = 'DCMM审计汇总' + suffix + '.xlsx'
    else:
        out_dir = default_out_dir
        excel_name = 'DCMM审计汇总.xlsx'
    return out_dir, excel_name


_source_dir_cache = {}


def find_source_dir(ent_name: str, ent_num: str = '', batch_dirs: list = None,
                     base_dir: str = None):
    """Find the enterprise's original report folder.

    Searches batch_dirs for a folder starting with the enterprise number.
    Returns (source_dir, source_link) where source_link points to the complete PDF.
    """
    if batch_dirs is None:
        batch_dirs = ['2、三级（第一天）']
    cache_key = f'{ent_num}_{ent_name}|{"|".join(batch_dirs)}'
    if cache_key in _source_dir_cache:
        return _source_dir_cache[cache_key]

    source_dir = None
    source_link = None

    for batch_dir_name in batch_dirs:
        bdir = batch_dir_name if os.path.isabs(batch_dir_name) else os.path.join(base_dir or '.', batch_dir_name)
        if not os.path.exists(bdir):
            continue
        for inst_dir_name in os.listdir(bdir):
            inst_dir_path = os.path.join(bdir, inst_dir_name)
            if not os.path.isdir(inst_dir_path) or inst_dir_name.startswith('.'):
                continue
            for sub_name in os.listdir(inst_dir_path):
                sub_path = os.path.join(inst_dir_path, sub_name)
                if not os.path.isdir(sub_path) or sub_name.startswith('.'):
                    continue
                if ent_num and sub_name.startswith(f'{ent_num}、'):
                    source_dir = sub_path
                    break
            if source_dir:
                break
        if source_dir:
            break

    if source_dir:
        complete_kw = ('完整', '完成版', 'complete')
        exclude_kw = ('评估申请书', '申请书', '承诺书', '检查表', '打分表')

        # Direct check
        for fname in os.listdir(source_dir):
            if fname.endswith('.pdf') and any(kw in fname for kw in complete_kw) and not fname.startswith('._'):
                source_link = os.path.join(source_dir, fname)
                break

        # Fallback: largest non-split PDF
        if not source_link:
            import fitz
            candidates = []
            for fname in os.listdir(source_dir):
                if not fname.endswith('.pdf') or fname.startswith('._') or fname.startswith('CESI完整版'):
                    continue
                if re.match(r'0?\d', fname) or any(kw in fname for kw in exclude_kw):
                    continue
                try:
                    d = fitz.open(os.path.join(source_dir, fname))
                    np = len(d)
                    d.close()
                except Exception:
                    continue
                candidates.append((np, os.path.join(source_dir, fname)))
            if candidates:
                candidates.sort(key=lambda x: -x[0])
                if candidates[0][0] >= 100:
                    source_link = candidates[0][1]

        # Subdirectory check
        if not source_link:
            for sub_name in os.listdir(source_dir):
                sub_path = os.path.join(source_dir, sub_name)
                if os.path.isdir(sub_path) and not sub_name.startswith('.'):
                    for fname in os.listdir(sub_path):
                        if fname.endswith('.pdf') and any(kw in fname for kw in complete_kw) and not fname.startswith('._'):
                            source_link = os.path.join(sub_path, fname)
                            break
                    if source_link:
                        break

        if not source_link:
            for r, _ds, fs in os.walk(source_dir):
                for fname in fs:
                    if fname.endswith('.docx') and not fname.startswith('~$') and not fname.startswith('._') and '评估报告' in fname:
                        source_link = os.path.join(r, fname)
                        break
                if source_link:
                    break
            if not source_link:
                source_link = source_dir

    _source_dir_cache[cache_key] = (source_dir, source_link)
    return source_dir, source_link
